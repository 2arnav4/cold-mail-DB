import json
import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def main():
    # 1. Read parsed PDF table data
    with open('/Users/arnavsingla/.gemini/antigravity/brain/9da516c5-a750-44c6-9750-85ca2711bd7c/scratch/pdfplumber_raw.json', 'r') as f:
        raw_table = json.load(f)
        
    # 2. Read resolved emails data
    with open('/Users/arnavsingla/.gemini/antigravity/brain/9da516c5-a750-44c6-9750-85ca2711bd7c/scratch/all_resolved_startups.json', 'r') as f:
        resolved_data = json.load(f)
        
    # Clean and filter table rows
    companies_data = {}
    for row in raw_table:
        if len(row) >= 8 and row[0].isdigit():
            row_num = int(row[0])
            if 2 <= row_num <= 102:
                companies_data[row[1].strip()] = {
                    'row_num': row_num,
                    'company': row[1].strip(),
                    'money_raised': row[2].strip(),
                    'round': row[3].strip(),
                    'month': row[4].strip(),
                    'sector': row[5].strip(),
                    'hq': row[6].strip(),
                    'founders_raw': row[7].strip()
                }
                
    print(f"Loaded {len(companies_data)} company profiles from PDF.")
    
    # 3. Connect to Turso local DB
    db_path = '/Users/arnavsingla/Desktop/Cold-Mail-DB/turso-full.db'
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Get current datetime
    now_str = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')
    
    db_inserted_companies = 0
    db_inserted_contacts = 0
    
    # 4. Prepare data for Excel and update DB
    excel_rows = []
    
    for comp_name, info in companies_data.items():
        # Get resolved details
        resolved = resolved_data.get(comp_name, {})
        domain = resolved.get('domain', f"{comp_name.lower().replace(' ', '')}.ai")
        emails = resolved.get('emails', [])
        
        # --- Update SQLite DB ---
        # A. Insert/Get Company
        # Try to find company by domain or exact name
        cursor.execute("SELECT id FROM companies WHERE LOWER(domain) = ? OR LOWER(name) = ?;", (domain.lower(), comp_name.lower()))
        co_row = cursor.fetchone()
        
        if co_row:
            company_id = co_row[0]
            # Update company details
            cursor.execute("""
                UPDATE companies 
                SET funding_stage = ?, industry = ?, batch = ?, domain = ?
                WHERE id = ?;
            """, (info['round'], info['sector'], info['month'], domain, company_id))
        else:
            cursor.execute("""
                INSERT INTO companies (domain, name, source, funding_stage, industry, description, created_at, batch)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?);
            """, (domain, comp_name, '51-Startups-raised-money', info['round'], info['sector'], f"Founders: {info['founders_raw']}", now_str, info['month']))
            company_id = cursor.lastrowid
            db_inserted_companies += 1
            
        # B. Insert/Get Contacts
        for email_info in emails:
            email = email_info['email']
            name = email_info['name']
            role = email_info['role']
            source = email_info['source']
            
            # Check if contact already exists
            cursor.execute("SELECT id FROM contacts WHERE LOWER(email) = ?;", (email.lower(),))
            contact_row = cursor.fetchone()
            
            email_confidence = 1.0 if 'DB:' in source else 0.8
            email_verified = 1 if 'DB:' in source else 0
            
            if not contact_row:
                cursor.execute("""
                    INSERT INTO contacts (company_id, name, role, email, email_confidence, email_verified, is_invalid, created_at, source, priority)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                """, (company_id, name, role, email, email_confidence, email_verified, 0, now_str, source, 1))
                db_inserted_contacts += 1
                
            # Add to Excel list
            excel_rows.append({
                'row_num': info['row_num'],
                'company': comp_name,
                'money_raised': info['money_raised'],
                'round': info['round'],
                'month': info['month'],
                'sector': info['sector'],
                'hq': info['hq'],
                'founders': info['founders_raw'],
                'domain': domain,
                'contact_name': name,
                'contact_role': role,
                'contact_email': email,
                'source': source
            })
            
    conn.commit()
    conn.close()
    
    print(f"Database updated. Inserted {db_inserted_companies} new companies and {db_inserted_contacts} new contacts.")
    
    # 5. Write to Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Startup Emails"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = [
        "Row #", "Company", "Money Raised", "Round", "Month (2026)", 
        "Sector", "HQ", "Founders", "Resolved Domain", 
        "Contact Name", "Contact Role", "Contact Email", "Email Source"
    ]
    
    # Font styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    
    # Fills
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Navy Blue
    
    # Borders
    thin_side = Side(style='thin', color='D9D9D9')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Write headers
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = cell_border
        
    # Sort excel rows by Row #
    excel_rows.sort(key=lambda x: x['row_num'])
    
    # Write data rows
    for r_data in excel_rows:
        row_values = [
            r_data['row_num'],
            r_data['company'],
            r_data['money_raised'],
            r_data['round'],
            r_data['month'],
            r_data['sector'],
            r_data['hq'],
            r_data['founders'],
            r_data['domain'],
            r_data['contact_name'],
            r_data['contact_role'],
            r_data['contact_email'],
            r_data['source']
        ]
        ws.append(row_values)
        
        row_idx = ws.max_row
        for col_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = cell_border
            
            # Alignments
            if col_idx in [1, 3, 4, 5]: # Row #, Money, Round, Month
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Save Workbook
    output_path = '/Users/arnavsingla/Desktop/Cold-Mail-DB/51-Startups-raised-money-emails.xlsx'
    wb.save(output_path)
    print(f"Excel file created successfully at: {output_path}")
    print(f"Total entries written to Excel: {len(excel_rows)}")

if __name__ == "__main__":
    main()
